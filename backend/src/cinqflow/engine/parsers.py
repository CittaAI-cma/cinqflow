"""CSV/XLSX bytes -> rows of strings. Typing is an inference made later, by the
profiler; the parser never coerces, so nothing is lost before it is observed."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field


class ParseError(Exception):
    pass


@dataclass
class ParsedFile:
    columns: list[str]
    rows: list[dict[str, str]]
    sheets: list[tuple[str, int]] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)


def _clean_header(raw: list[str]) -> list[str]:
    columns, seen = [], {}
    for i, name in enumerate(raw):
        col = (name or "").strip() or f"column_{i + 1}"
        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
        columns.append(col)
    return columns


def parse_csv(content: bytes) -> ParsedFile:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    if not text.strip():
        raise ParseError("file is empty")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header = next(reader)
    except StopIteration as exc:
        raise ParseError("file has no header row") from exc

    columns = _clean_header(header)
    rows: list[dict[str, str]] = []
    for raw in reader:
        if not any((cell or "").strip() for cell in raw):
            continue
        padded = list(raw) + [""] * (len(columns) - len(raw))
        rows.append({col: (padded[i] or "").strip() for i, col in enumerate(columns)})
    return ParsedFile(columns=columns, rows=rows, sheets=[("csv", len(rows))])


def parse_xlsx(content: bytes) -> ParsedFile:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ParseError("openpyxl unavailable") from exc

    try:
        book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"cannot read xlsx: {exc}") from exc

    sheet = book[book.sheetnames[0]]
    it = sheet.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration as exc:
        raise ParseError("sheet is empty") from exc

    columns = _clean_header(["" if c is None else str(c) for c in header])
    rows: list[dict[str, str]] = []
    for raw in it:
        if raw is None or not any(c is not None and str(c).strip() for c in raw):
            continue
        padded = list(raw) + [None] * (len(columns) - len(raw))
        rows.append(
            {
                col: ("" if padded[i] is None else str(padded[i]).strip())
                for i, col in enumerate(columns)
            }
        )

    sheets = [(name, 0) for name in book.sheetnames]
    sheets[0] = (book.sheetnames[0], len(rows))
    book.close()
    return ParsedFile(columns=columns, rows=rows, sheets=sheets)


def parse(content: bytes, file_type: str) -> ParsedFile:
    if file_type == "csv":
        return parse_csv(content)
    if file_type == "xlsx":
        return parse_xlsx(content)
    raise ParseError(f"unsupported file type: {file_type}")
