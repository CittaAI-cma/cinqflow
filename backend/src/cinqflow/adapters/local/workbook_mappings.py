"""Read the client's real source-to-Silver-Raw mapping workbooks. NOT a fixture.

    "golden sets are harvested from artifacts the programme already produced,
     never written for the occasion. A test set invented alongside the thing it
     tests measures nothing."
    — memory/05-ground-truth/03-golden-sets.md

The source is `clientdata/Uploads/Claims Mapping/*.xlsx`: 188 MAPPED rows on the
Fidelis sheet alone, each a decision one of the client's analysts made about
which Silver-Raw column a payer's field populates. They are the answer key for
CF-V1-E6-02's blind re-derivation gate, and the reason that gate is a
benchmark rather than an approximation.

THE WORKBOOKS ALSO CARRY THE TARGET MODEL. Every MAPPED row names an `SR Table`
and an `SR Column`, and most carry an `SR Description`. Together those are the
canonical vocabulary a BA would have been shown — so the eval can build the
target model from the same artefact the answer key came from, rather than from
one invented for it.

This lives in `adapters/` because it opens a file. `core/mapping` holds the
shapes and every question asked of them, and knows nothing about openpyxl,
paths or sheets.
"""

from __future__ import annotations

import re
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from cinqflow.adapters.local.workbook_glossary import WorkbookError

#: The status the client's analysts write when a source field HAS a target.
#: Their vocabulary also includes `AUDIT` (the pipeline supplies it),
#: `NO MAP` / `NO DL COL` (nothing populates it, with a reason) and `NEW`.
#: Only MAPPED rows are answers; the rest are decisions of other kinds.
MAPPED = "MAPPED"


@dataclass(frozen=True)
class GoldenMapping:
    """One row of a client mapping workbook: a decision, with its provenance."""

    source_field: str
    target_entity: str
    target_field: str
    notes: str = ""
    description: str = ""
    status: str = MAPPED
    source_row: int | None = None

    @property
    def address(self) -> str:
        return f"{self.target_entity}.{self.target_field}"


def _rows(path: Path, sheet: str) -> Iterator[tuple[int, dict[str, Any]]]:
    """Rows keyed BY HEADER NAME, with the header row found rather than assumed.

    Found rather than assumed because these workbooks have a blank first row
    (and sometimes four), and a loader that hardcoded row 2 would read the
    banner as its headers on the next workbook somebody hands us.
    """
    try:
        import openpyxl
    except ImportError as missing:  # pragma: no cover - environment, not logic
        raise WorkbookError(
            "reading the client workbooks needs openpyxl — install requirements/formats.txt"
        ) from missing

    if not path.exists():
        raise WorkbookError(f"no workbook at {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        raise WorkbookError(
            f"{path.name} has no {sheet!r} sheet — found: {', '.join(workbook.sheetnames)}"
        )
    rows = list(workbook[sheet].iter_rows(values_only=True))
    workbook.close()

    header_at = _header_row(rows)
    if header_at is None:
        raise WorkbookError(f"{path.name}/{sheet}: no row carries an 'SR Column' header")
    headers = [str(h).strip() if h is not None else "" for h in rows[header_at]]
    for index, row in enumerate(rows[header_at + 1 :], start=header_at + 2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        yield index, dict(zip(headers, row, strict=False))


def _header_row(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows[:8]):
        labels = {str(cell).strip().lower() for cell in row if cell is not None}
        if "sr column" in labels and "sr table" in labels:
            return index
    return None


#: The source-field column is spelled per payer — `Fidelis Source Field`,
#: `CCLF Field`. Matched by shape rather than listed, so a new payer's workbook
#: reads without a code change; the SR side is spelled consistently and is not.
_SOURCE_HEADER = re.compile(r"(source field|^.* field$)", re.IGNORECASE)


def load_mappings(
    path: str | Path, sheet: str, *, expect: int | None = None
) -> tuple[GoldenMapping, ...]:
    """The MAPPED rows of one sheet, as decisions.

    `expect` is checked and REPORTED rather than assumed: a golden set that
    silently loaded half its rows would report a flattering pass on an exam
    with most of the questions missing.
    """
    source = Path(path)
    found: list[GoldenMapping] = []
    for index, row in _rows(source, sheet):
        headers = list(row)
        source_key = next((h for h in headers if _SOURCE_HEADER.search(h)), "")
        source_field = str(row.get(source_key) or "").strip()
        entity = str(row.get("SR Table") or "").strip()
        field = str(row.get("SR Column") or "").strip()
        status = str(row.get("Status") or "").strip()
        if status != MAPPED or not (source_field and entity and field):
            continue
        found.append(
            GoldenMapping(
                source_field=source_field,
                target_entity=entity,
                target_field=field,
                notes=str(row.get("Transform / Mapping Notes") or row.get("Mapping Notes") or ""),
                description=str(row.get("SR Description") or ""),
                status=status,
                source_row=index,
            )
        )
    if expect is not None and len(found) != expect:
        raise WorkbookError(
            f"{source.name}/{sheet}: expected {expect} MAPPED rows, read {len(found)}. The "
            "programme's counted figure and the workbook disagree — record the discrepancy "
            "rather than silently picking a side."
        )
    return tuple(found)


def distinct_pairs(mappings: tuple[GoldenMapping, ...]) -> tuple[GoldenMapping, ...]:
    """One row per (source field, target) pair, first occurrence kept.

    The client's sheets repeat a pairing across claim-type scopes — `claim_id`
    appears once per scope — and grading the same decision five times would
    weight it five times.
    """
    seen: dict[tuple[str, str], GoldenMapping] = {}
    for row in mappings:
        seen.setdefault((row.source_field.lower(), row.address.lower()), row)
    return tuple(seen.values())


#: A trailing ordinal on a repeated column: `diagnosis_1`, `poa_7`,
#: `other_procedure_10`. Stripped ANYWHERE in the name, not only at the end,
#: because the client also writes `diagnosis_10_poa`.
_ORDINAL = re.compile(r"_\d+(?=_|$)")


def distinct_decisions(mappings: tuple[GoldenMapping, ...]) -> tuple[GoldenMapping, ...]:
    """One row per distinct DECISION — the exam a gate should be set on.

    75 of the Fidelis sheet's 188 pairs are `diagnosis_1..11 ->
    claim_diagnosis.source_diagnosis_code` and `poa_1..11 ->
    claim_diagnosis.poa_flag`: eleven copies of two decisions, because the
    payer sends the same concept in numbered columns and the target unpivots
    them into rows. An agent that gets `diagnosis_1` right gets all eleven
    right for free.

    Grading over pairs would let that one easy family carry 40% of the score,
    and a gate passed that way would say nothing about the columns that are
    actually hard. So the ordinal suffix is stripped and the family collapses
    to the decisions it is.

    Both numbers are worth reporting — `distinct_pairs` is the work saved, and
    this is the difficulty faced.
    """
    seen: dict[tuple[str, str], GoldenMapping] = {}
    for row in mappings:
        stem = _ORDINAL.sub("", row.source_field.lower())
        seen.setdefault((stem, row.address.lower()), row)
    return tuple(seen.values())
