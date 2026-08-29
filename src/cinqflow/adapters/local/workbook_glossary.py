"""Read the client's real glossary and DQ-rule sheets. NOT a fixture.

    "golden sets are harvested from artifacts the programme already produced,
     never written for the occasion. A test set invented alongside the thing it
     tests measures nothing."
    — memory/05-ground-truth/03-golden-sets.md

The source is `thiers/Uploads/2-Design/Data lake data model.xlsx`: 171 business
glossary rows and 110 data-quality rules, written by the client's own analysts
before this platform existed. They serve twice (ADR-0007) — as the exam for the
PHI-detection and NL-rule gates, and as K2 grounding once E16-04 embeds them.

This lives in `adapters/` because it opens a file. `core/registry/glossary.py`
holds the shapes and every question asked of them, and knows nothing about
openpyxl, paths or sheets.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

from cinqflow.core.registry.glossary import Glossary, GlossaryTerm

GLOSSARY_SHEET = "Business Glossary"
DQ_RULES_SHEET = "Data Quality Rules"

#: What the programme has counted and documented. A load that returns a
#: different number means the source changed, and that is a fact to surface
#: rather than a discrepancy to absorb silently.
EXPECTED_TERMS = 171
EXPECTED_RULES = 110


class WorkbookError(RuntimeError):
    """The source workbook is not the shape the platform was told it is."""


def _rows(path: Path, sheet: str) -> Iterator[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as missing:  # pragma: no cover - environment, not logic
        raise WorkbookError(
            "reading the client workbooks needs openpyxl — install requirements/formats.txt"
        ) from missing

    if not path.exists():
        raise WorkbookError(f"no workbook at {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        raise WorkbookError(
            f"{path.name} has no {sheet!r} sheet — found: {', '.join(workbook.sheetnames)}"
        )
    worksheet = workbook[sheet]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(iterator)]
    except StopIteration:
        raise WorkbookError(f"{sheet!r} is empty") from None
    # Keyed BY HEADER NAME, so a column reorder in the source cannot silently
    # shift a PHI flag onto the wrong term.
    for row in iterator:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        yield dict(zip(headers, row, strict=False))


def load_glossary(path: str | Path, *, expect: int | None = EXPECTED_TERMS) -> Glossary:
    """The 171 terms, with their PHI flags and synonym sets.

    `expect` is checked and REPORTED rather than assumed: a seed that silently
    loaded 140 terms would weaken the 100%-recall PHI gate by 31 terms and
    nothing would say so.
    """
    source = Path(path)
    terms = tuple(
        GlossaryTerm.from_row(row, source_row=index)
        for index, row in enumerate(_rows(source, GLOSSARY_SHEET), start=2)
        if str(row.get("Glossary_ID") or "").strip()
    )
    if expect is not None and len(terms) != expect:
        raise WorkbookError(
            f"{source.name}: expected {expect} glossary terms, read {len(terms)}. The "
            "programme's counted figure and the workbook disagree — record the discrepancy "
            "rather than silently picking a side."
        )
    return Glossary(terms=terms)


def load_dq_rule_rows(
    path: str | Path, *, expect: int | None = EXPECTED_RULES
) -> tuple[dict[str, Any], ...]:
    """The 110 legacy rules, raw.

    Each already pairs a natural-language description with executable SQL and a
    glossary link — which is exactly the shape CF-V1-E7-01's agent must
    produce, so the exam is a re-derivation benchmark rather than an
    approximation. Shaping them into governed DQ rules is E7-03's job; this
    returns the rows so the harness and the seeder can each do their own thing.
    """
    source = Path(path)
    rows = tuple(
        row for row in _rows(source, DQ_RULES_SHEET) if str(row.get("DQ_Rule_ID") or "").strip()
    )
    if expect is not None and len(rows) != expect:
        raise WorkbookError(f"{source.name}: expected {expect} DQ rules, read {len(rows)}")
    return rows
